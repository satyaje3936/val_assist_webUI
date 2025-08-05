import openpyxl
import os
import csv
import re
import json
import re
import json
import pandas as pd
import math

file_path = r"C:\PythonSv\graniterapidsd\users\kvembar\FCO_Framework_Development" 
output_path = r"C:\PythonSv\graniterapidsd\users\kvembar\output_file"
#output_excel_jason_cs_mapping_path= r"C:\PythonSv\graniterapidsd\users\kvembar\Fuse_check\XCC\output_file\json_cs_mapping.xlsx"

def read_excel_fuse_file(file_path,sku):
    print("In the called function read_excel_file")
    from openpyxl import load_workbook
    
    new_file_path = os.path.join(file_path,"FuseReport_XCP_GNRD_A0_17_2.xlsx")
    #new_file_path = os.path.join(file_path,"FuseReport.xlsx")
    print(new_file_path)
    workbook = load_workbook(new_file_path)
    
    fuse_list = []
    regdict = {}
    fuse_dict = {}
    
    #sheets = [ "CPU_GRANITE_RAPIDS_COMPUTE_DIE", "IO_Device_Granite_Rapids_D_IO_","DIE_PIROM"]
    sheets = [ "CPU_GRANITE_RAPIDS_COMPUTE_DIE"]
    
    for sheet_name in sheets:
        if sheet_name in workbook.sheetnames:
            work_sheet = workbook[sheet_name]
            print(f"Opened sheet: {work_sheet.title}")
    
    
            # Accessing the 7th row to get headers
            header_row = [cell.value for cell in work_sheet[7]]   # accessing the 7th row, It extracts the values of each cell in this row into a list.
            print(f"Printing the header row:{header_row}")
            
            # Find column indices based on headers
            fuse_name_col = header_row.index("Fuse Name")+1
            print(f"Fuse Name Column: {fuse_name_col}")    
            
            die_name_index = header_row.index("Register") + 1 
            
            if sku not in header_row:
                print(f"⚠️ SKU '{sku}' not found in sheet '{sheet_name}' — skipping.")
                continue
                
            qdf_fuse_value = header_row.index(sku) + 1 
            
            comment_col = header_row.index("Comment") + 1  
            
            die_name = [cell.value for cell in work_sheet[8]]
            die_name = die_name[die_name_index -1]
            print(f"Die Name: {die_name}\n")
            
            max_value_row = work_sheet.max_row
            print( f"maximum rows:{max_value_row}")
            
            #regdict = {}
            
            i = 0
            for row in work_sheet.iter_rows(min_row = 8, min_col = 1, max_row = max_value_row):
                reglist = []
                for cell in row:
                    reglist.append(cell.value)
                regdict[i] = reglist
                i = i+1
            #print(f"regdict: {regdict} \n")
    
        else:
            print(f"Sheet '{sheet_name}' does not exist.")
            break
     

        

        for row in work_sheet.iter_rows(min_row=8): # Start from row 8 to skip headers
            fuse_cell = row[fuse_name_col - 1].value  # Get the "Fuse name" value
            fuse_name = str(fuse_cell).strip().split('/')[-1].lower()
            
            value_cell = row[qdf_fuse_value - 1]
            qdf_fuse_value_hex = value_cell.value if value_cell else None
            
            if qdf_fuse_value_hex == 'm':
                qdf_fuse_value_dec = 'dynamic value (m)'
            elif isinstance(qdf_fuse_value_hex, str) and 'h' in qdf_fuse_value_hex:
                try:
                    qdf_fuse_value_dec = qdf_fuse_value_hex 
                    # hex_value = qdf_fuse_value_hex.split('h')[-1]
                    # qdf_fuse_value_dec = int(hex_value, 16)
                except ValueError:
                    qdf_fuse_value_dec = 'invalid hex'
            else:
                qdf_fuse_value_dec = qdf_fuse_value_hex  # Could be already int or None
            
            #fuse_list.append([fuse_name, qdf_fuse_value_dec])
            comment_value = row[comment_col - 1].value if comment_col <= len(row) else ""
            fuse_list.append([fuse_name, qdf_fuse_value_dec, comment_value])
            
            #fuse_dict = {fuse_name: value for fuse_name, value in fuse_list}
            fuse_dict = {fuse_name: {"value": value, "comment": comment} for fuse_name, value, comment in fuse_list}
    
    return regdict, fuse_dict
    
# def extract_fuse_values_from_sheet(work_sheet, fuse_name_col, qdf_fuse_value_col):
    # fuse_list = []

    # for row in work_sheet.iter_rows(min_row=8):
        # fuse_cell = row[fuse_name_col - 1].value
        # if not fuse_cell:
            # continue
        
        # fuse_name = str(fuse_cell).strip().split('/')[-1].lower()

        # value_cell = row[qdf_fuse_value_col - 1]
        # qdf_fuse_value_hex = value_cell.value if value_cell else None

        # # Convert value
        # if qdf_fuse_value_hex == 'm':
            # qdf_fuse_value_dec = 'dynamic value (m)'
        # elif isinstance(qdf_fuse_value_hex, str) and 'h' in qdf_fuse_value_hex:
            # try:
                # hex_value = qdf_fuse_value_hex.split('h')[-1]
                # qdf_fuse_value_dec = int(hex_value, 16)
            # except ValueError:
                # qdf_fuse_value_dec = 'invalid hex'
        # else:
            # qdf_fuse_value_dec = qdf_fuse_value_hex

        # fuse_list.append([fuse_name, qdf_fuse_value_dec])

    # return fuse_list

            
# def read_excel_lava_fuse_file(file_path):
    # print("In the called function read_excel_lava_fuse_file")
    
    # from openpyxl import load_workbook 
    
    # lava_fuse_file_path = os.path.join(file_path,"y25ww14_Fuse File Equation Extract (GNRD_B1_11_0).xlsx")
    
    # workbook = load_workbook(lava_fuse_file_path)
    # worksheet = workbook.active
    # print(f"Opened sheet: {worksheet.title}")
    
    # header_row = [cell.value for cell in worksheet[1]] 
    # print(f"Printing Header row: {header_row}")
    
    # fuse_name_col = header_row.index("Fuses")+1
    # #print(fuse_name_col)
    # fuse_equation_col = header_row.index("Utilization Methods")+1
    # #print(fuse_equation_col)
    # lira_attributes_col=header_row.index("Line Item Attributes (Equation)")+1 
    # #print(lira_attributes_col)
     
    # max_value_row = worksheet.max_row
    # print( f"maximum rows:{max_value_row}") 
    
    # res_dict = {}
    # i=0
    
    # for i, row in enumerate(worksheet.iter_rows(min_row=2, min_col=1, max_row=max_value_row, values_only=True), start=1):
        # res_dict[i] = {
            # "Fuses": row[fuse_name_col - 1],
            # "Utilization Methods": row[fuse_equation_col - 1],
            # "Line Item Attributes (Equation)": row[lira_attributes_col - 1]
        # }
    # print(f"res_dict: {res_dict} \n")
    
    # return res_dict
    

# def read_csv_lava_file(file_path):
    
    # print("In the called function read_csv_lava_file")
    
    # csv_file_path = os.path.join(file_path,"ES1 LAVA export ww16.3.csv")
    # print(csv_file_path)
    
    # # Initialize an empty list to store CSV data
    # csv_data_list = []
    
    # # Open the CSV file and read its contents
    # try:
        # with open(csv_file_path, newline='', encoding='utf-8') as file:
             # csv_reader = csv.reader(file)
        
        # for row in csv_reader:
            # csv_data_list.append(row)
        
        # print(csv_data_list)
        
    # except Exception as e:
        # print(f"An error occurred: {e}")
        
    # # Return the list containing CSV data    
    # return csv_data_list
    
    
def read_lava_jason_extract(file_path):
    print("In the called function read_lava_json_file")
    # Path to your JSON file
    #json_file_path = "C:\pythonsv\graniterapidsd\users\kvembar\Fuse_check\XCC\Q7YT.json"
    json_file_path = os.path.join(file_path,"Q7YT_ORG.json")
    print(json_file_path)  
    
    # Open the JSON file and load its contents into a dictionary
    try:
        with open(json_file_path, 'r') as file:
            data = json.load(file)
            #print(data)
            return data
            
    except FileNotFoundError:
        print(f"Error: The file {json_file_path} does not exist.")
    
    #exception that occurs when there is an error in decoding JSON data
    except json.JSONDecodeError:
        print("Error: Failed to decode JSON. Please check the file format.")
    
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

def read_equations_cs_file_extract(file_path):
    print("In the called function read_equations_cs_file_extract")
    cs_file_path = os.path.join(file_path,"CPU_GRANITE_RAPIDS_COMPUTE_DIE_M_Rule_org.cs")
    print(cs_file_path)

    #Loads the entire .cs file into a string.
    with open(cs_file_path, 'r', encoding='utf-8', errors='ignore') as file:
        content = file.read()
    
    # LIRAMappingWrapper(to match or identify calls to a function or method named LIRAMappingWrapper)
    lira_pattern = re.compile(r'LIRAMappingWrapper\(\s*lineItem\.([\w\d_]+)\.Value\s*,\s*(Fuses\.[^,]+)\s*,\s*new Dictionary<string, string>\s*{(.*?)}\s*\);',re.DOTALL)

    #Directfuse mapping (to match or identify calls to a function or method named fuse)
    fuse_equation_pattern = re.compile(r'(Fuses\.[\w\.]+)\.Fuse(SetValue|BinaryValue)\s*=\s*(.*?);',re.DOTALL)
    

    equation_results = []

    # Extract LIRA mappings
    for match in lira_pattern.finditer(content):                             #Iterates over all matches of the lira_pattern in the file content.
        lira_attr = match.group(1).strip()
        fuse_name = match.group(2).strip()
        equation = match.group(3).strip().replace("\n", "").replace(" ", "")
        equation_results.append({
            'Source': 'LIRAWrapper',
            'LIRA Attribute': lira_attr,
            'Fuse Name': fuse_name,
            'Equation': equation
        })

    #Extract fuse set equations
    for match in fuse_equation_pattern.finditer(content):                   #Iterates over all matches of the fuse_equation_pattern in the file content
        fuse_name = match.group(1).strip()
        equation = match.group(3).strip()
        #working code 
        # lira_attr_match = re.search(r'lineItem\.([\w\d_]+)\.Value', equation)
        # lira_attr = lira_attr_match.group(1) if lira_attr_match else ''
        
        lira_attr_matches = re.findall(r'lineItem\.([\w\d_]+)\.Value', equation)              #re.findall gets all LIRA attribute references from the equation.
        #lira_attr = sorted(set(lira_attr_matches)) if lira_attr_matches else [] 
        lira_attr = ', '.join(sorted(set(lira_attr_matches))) if lira_attr_matches else 'NA'
        #print(f"kavya Fuse: {fuse_name}, LIRA Attributes: {lira_attr}, Equation: {equation}")
        equation_results.append({
            'Source': 'DirectAssignment',
            'LIRA Attribute': lira_attr,
            'Fuse Name': fuse_name,
            'Equation': equation
        })
        
    return equation_results
        
def export_cs_file_extract_to_excel(equation_results, output_path):
    df = pd.DataFrame(equation_results)
    output_file = os.path.join(output_path, 'fuse_extraction_results_file1.xlsx')
    df.to_excel(output_file, index=False)
    #print(f"Equation Results exported to: {output_file}")


def map_json_with_cs(lira_json, equation_results, output_path):    
    
    mapped_result = []
    
    #print("Kavya LIRA JSON :", lira_json )

    #print("\n KavyaResults:", equation_results)
    
    # Step 1: Extract attributeName and attributeValue from collections
    attribute_data_collections = {}
    for collection in lira_json.get("collections", []):
        for attribute in collection.get("attributes",[]):
            Name = attribute.get("attributeName")
            Value = attribute.get("attributeValue")
            if Name:
               attribute_data_collections[Name.strip()] = Value

    print(f" \n attribute_data_collections: {attribute_data_collections} \n")

    for map_entry in equation_results:
        print(f"map_entry: {map_entry}")
        lira_attrs = [attr.strip() for attr in map_entry["LIRA Attribute"].split(',')]
        # Check if all attributes in lira_attrs are present in attribute_data_collections
        if all(attr.lower() in [a.lower() for a in attribute_data_collections.keys()] for attr in lira_attrs):
            # If multiple attributes, put both names in LIRA Attribute field and their values
            lira_values = [attribute_data_collections.get(attr, "") for attr in lira_attrs]
            mapped_result.append({
                'Source': map_entry['Source'],
                'LIRA Attribute': ', '.join(lira_attrs),
                'LIRA Value': ', '.join(lira_values),
                'Fuse Name': map_entry['Fuse Name'],
                'Equation': map_entry['Equation']
            })
            matched = True
        else:
            # Direct assignment case: no LIRA attribute match
            mapped_result.append({
            'Source': map_entry['Source'],
            'LIRA Attribute': 'NA',
            'LIRA Value': 'NA',
            'Fuse Name': map_entry['Fuse Name'],
            'Equation': map_entry['Equation']
            })

        
        # if not matched:
        #     print(f"Kavya No match found for {lira_attrs}")
         
    # # Step 2:Mapping the LIRA attribute Name and Value in json file and .cs file
    # for attributeName, attributeValue in attribute_data_collections.items():            
    #     #print(f"Kavya Processing LIRA Attribute: {attributeName}, Value: {attributeValue}") 
    #     matched = False
    #     for map_entry in equation_results:
    #         #print("Kavya Present Map Entry:", map_entry)
    #         #print(f"Kavya Match found for in json {attributeName}")
    #         #print(f"Kavya Match found for in cs file {map_entry['LIRA Attribute']}")
            
    #         lira_attrs = map_entry["LIRA Attribute"]   #retrieving the value stored in the "LIRA Attribute" field from the map_entry dictionary, which can be either a list of strings (e.g., ['_0_DTS_Max', '_0_DTS_Cal_Guardband']) or a single string (e.g., '_0_DTS_Max').
    #         if isinstance(lira_attrs, str):
    #             lira_attrs = [lira_attrs]  # fallback for existing rows

    #         print(f"map_entry: {map_entry}\n")
    #         #if attributeName.strip().lower() in [attr.strip().lower() for attr in map_entry["LIRA Attribute"]]: #This checks if the current attributeName (from your LIRA JSON) matches any of the attributes in the list
    #         if attributeName.strip().lower()  == map_entry["LIRA Attribute"].strip().lower(): #This checks if the current attributeName (from your LIRA JSON) matches any of the attributes in the list
    #             #print(f"Kavya Match found for {attributeName}")
    #             #print(f"Kavya Appending to result: {attributeName} = {attributeValue}")
    #             mapped_result.append({
    #                 'Source':map_entry['Source'],
    #                 'LIRA Attribute': attributeName,
    #                 'LIRA Value': attributeValue,
    #                 'Fuse Name': map_entry['Fuse Name'],
    #                 'Equation': map_entry['Equation']
    #             })
    #             matched = True
    
    #     if not matched:
    #         print(f"Kavya No match found for {attributeName}")
            
    #print("Kavya Mapped Results:", mapped_result)
    return mapped_result
    
    
def export_map_json_cs_file_extract_to_excel(mapped_result, output_path):
    df = pd.DataFrame(mapped_result)
    output_file = os.path.join(output_path, 'fuse_extraction_results_with_LAVA_values_file2.xlsx')
    df.to_excel(output_file, index=False)
    print(f"Results with LAVA values exported to: {output_file}")   


def map_fuse_report_with_cs(fuse_data, mapped_result, sku):   
    mapped_fuses = []

    for cs_entry in mapped_result:
        fuse_name = cs_entry.get("Fuse Name", "").strip().split('.')[-1].lower()
        if fuse_name in fuse_data:
            actual_fuse_value = fuse_data[fuse_name].get("value", "Not found")
            comment = fuse_data[fuse_name].get("comment", "Not found")
            if not comment or str(comment).strip() == "":
                comment = "Not found"
            mapped_fuses.append({
                "Source": cs_entry.get("Source", ""),
                "LIRA Attribute": cs_entry.get("LIRA Attribute", ""),
                "LIRA Value": cs_entry.get("LIRA Value", ""),
                "Fuse Name": fuse_name,
                "Actual Value from Report": actual_fuse_value,
                "HSD Info": comment,
                "Equation": cs_entry.get("Equation", "")
            })
        else:
            fuse_name = "Not found"
            actual_fuse_value = "Not found"
            comment = "Not found"
            
            mapped_fuses.append({
                "Source": cs_entry.get("Source", ""),
                "LIRA Attribute": cs_entry.get("LIRA Attribute", ""),
                "LIRA Value": cs_entry.get("LIRA Value", ""),
                "Fuse Name": fuse_name,
                "Actual Value from Report": actual_fuse_value,
                "HSD Info": comment,
                "Equation": cs_entry.get("Equation", "")
            })
            print(f"Kavya Fuse {fuse_name} not found in report")
            
    return mapped_fuses
    
def export_fuse_mapping_to_excel(fuse_mapped_result, output_path,sku):
    df = pd.DataFrame(fuse_mapped_result)
    output_file = os.path.join(output_path, f'{sku}_fuse_mapping_file3.xlsx')
    df.to_excel(output_file, index=False)
    print(f"Fuse name from .cs mapped result with fuse report exported to: {output_file}")
 


def evaluate_equations(equation_results, lira_json, fuse_dict_current, output_path, sku):
    calculated_equation_output_data = []

    # Step 1: Build a dictionary of LIRA attribute values from the JSON
    attribute_map = {}
    for collection in lira_json.get("collections", []):
        for attribute in collection.get("attributes", []):
            name = attribute.get("attributeName")
            value = attribute.get("attributeValue")
            if name:
                attribute_map[name.strip()] = value

 

    # Step 2: Process each equation
    for item in equation_results:
        source = item.get("Source", "Unknown")
        lira_attrs = item.get("LIRA Attribute", "Unknown")  # might contain multiple attrs
        lira_value = item.get("LIRA Value", "")
        fuse_name = item.get("Fuse Name", "Unknown").strip().split('.')[-1].lower()
        #equation = item.get("Equation", "")
        equation = item.get("Raw Equation", "") or item.get("Equation", "")
        #actual_value = fuse_dict_current.get(fuse_name, "N/A")
        actual_entry = fuse_dict_current.get(fuse_name, {})
        actual_value = actual_entry.get("value", "N/A") if isinstance(actual_entry, dict) else actual_entry
        hsd_info = actual_entry.get("comment", "") if isinstance(actual_entry, dict) else "",
        processed_eqn = equation
        
        # Ensure lira_attrs is a list
        if isinstance(lira_attrs, str):
            lira_attrs = [lira_attrs]

        # Clean equation for regex scan
        cleaned_eqn = str(equation).replace('"', '').replace("'", "")
        
        # Extract all LIRA attribute references case-insensitively
        matched_attrs = re.findall(r'lineitem\.([A-Za-z0-9_]+)\.value', cleaned_eqn, flags=re.IGNORECASE)
        
        # Debug print for verification
        if not matched_attrs:
            #print(f"[WARN] No LIRA attributes found in: {fuse_name} | Equation: {raw_eqn}")
            print(f"[WARN] No LIRA attributes found in: {fuse_name} | Equation: {equation}")
        else:
            print(f"[INFO] Found LIRA attributes in {fuse_name}: {matched_attrs}")
        
        # Step 3: Replace LIRA attribute placeholders in equation with actual values
        #matched_attrs = re.findall(r'lineItem\.([A-Za-z0-9_]+)\.Value', processed_eqn)
        #lira_attrs = matched_attrs  # ← this ensures correct attributes are captured   #added 1
        lira_values_used = []
        
        # Fallback if LIRA attribute field was not correctly populated
        if not matched_attrs:
            lira_attrs = ["NA"]
            lira_values_used = ["NA"]
        else: 
            # Use matched_attrs as LIRA attributes
            lira_attrs = matched_attrs
            lira_values_used = []
            for attr in matched_attrs:
                val = attribute_map.get(attr, "0")
                lira_values_used.append(f"{attr}: {val}")
                processed_eqn = processed_eqn.replace(f"lineItem.{attr}.Value", str(val))
            
        def convert_ternary(expr):
            expr = expr.replace("&&", "and").replace("||", "or")

            pattern = re.compile(r'([^\?:]+?)\?([^\?:]+?):([^\?:]+)')
            
            while '?' in expr and ':' in expr:
                    match = pattern.search(expr)
                    if not match:
                        break
                    condition = match.group(1).strip()
                    if_true = match.group(2).strip()
                    if_false = match.group(3).strip()
                    python_expr = f"({if_true} if {condition} else {if_false})"
                    expr = expr[:match.start()] + python_expr + expr[match.end():]
            return expr
   
        processed_eqn = convert_ternary(processed_eqn)
        # # Step 4: Translate .cs syntax to Python
        processed_eqn = processed_eqn.replace("Math.Round", "round") 
        processed_eqn = processed_eqn.replace("Math.Floor", "math.floor")        
        # Replace dec2hex with hex(int(round(...)))
        processed_eqn = re.sub(r'dec2hex\s*\(\s*round\((.*?)\)\s*\)', r'hex(int(round(\1)))[2:].upper()', processed_eqn)
        #processed_eqn = re.sub(r'dec2hex\s*\(\s*round\((.*?)\)\s*\)', lambda m: f"hex(int(round(float({m.group(1)}))))[2:].upper()", processed_eqn)
        processed_eqn = re.sub(r'dec2hex\s*\(\s*(.*?)\s*\)', r'hex(int(round(\1)))[2:].upper()', processed_eqn)
        #processed_eqn = re.sub(r'dec2hex\s*\(\s*(.*?)\s*\)', lambda m: f"hex(int(round(float({m.group(1)}))))[2:].upper()", processed_eqn)
        # Convert 6'hA to 0xA format
        processed_eqn = re.sub(r"(\d+)'h([0-9a-fA-F]+)", r"0x\2", processed_eqn)
        processed_eqn = re.sub(r'Convert\.ToDouble\((.*?)\)', r'float(\1)', processed_eqn)
        processed_eqn = re.sub(r'"?(\d+)\'h([0-9a-fA-F]+)"?', lambda m: hex(int(m.group(2), 16)), processed_eqn)
        processed_eqn = re.sub(r'\(?Convert\.ToDouble\((.*?)\)\)?', lambda m: f'float({m.group(1)})', processed_eqn)
        #processed_eqn = re.sub(r'dec2hex\s*\(\s*math\.floor\((.*?)\)\s*\)', r'hex(int(math.floor(\1)))[2:].upper()', processed_eqn)
        #processed_eqn = re.sub(r'dec2hex\s*\(\s*round\(math\.floor\((.*?)\)\)\s*\)', r'hex(int(math.floor(\1)))[2:].upper()', processed_eqn)
        processed_eqn = re.sub(r'dec2hex\s*\(\s*math\.floor\((.*?)\)\s*\)', r"hex(int(math.floor(\1))).replace('0x','').upper()", processed_eqn)
        processed_eqn = re.sub(r'dec2hex\s*\(\s*round\(math\.floor\((.*?)\)\)\s*\)', r"hex(int(math.floor(\1))).replace('0x','').upper()", processed_eqn)
        processed_eqn = re.sub(r'dec2hex\s*\(\s*round\((.*?)\)\s*\)',r'str(hex(int(round(\1)))).replace("0x", "").upper()',processed_eqn)

        # Step 5: Evaluate
        try:
            if processed_eqn.strip().startswith("{") and ':' in processed_eqn:
                dict_obj = eval(processed_eqn)
                key = str(lira_value).strip().upper()
                result = dict_obj.get(key, f"Key '{key}' not found")
            elif not re.search(r'[+\-*/<>=()]', processed_eqn) and not processed_eqn.strip().startswith(('lineItem', 'dec2hex')):
                result = processed_eqn
            else:
                result = eval(processed_eqn)
        except Exception as e:
            result = f"ERROR: {str(e)}"

        # Step 6: Append to results
        # if source == "DirectAssignment":
            # # Try to evaluate equation like "3'h3"
            # try:
                # evaluated_value = eval(processed_eqn)
            # except Exception as e:
                # evaluated_value = f"ERROR: {str(e)}"
        # #if not matched_attrs and not equation.strip():
            # # No LIRA attr, no equation => direct assignment case
            # calculated_equation_output_data.append({
                # "Source": source,
                # "LIRA Attribute": "No LIRA Attribute (DirectAssignment)",
                # "LIRA Value": "No LIRA Value (DirectAssignment)",
                # "Fuse Name": fuse_name,
                # "Actual Value from Report": actual_value,
                # "Equation": "",
                # "Calculated Value": actual_value
            # })
        # else:
            # calculated_equation_output_data.append({
                # "Source": source,
                # "LIRA Attribute": ', '.join(lira_attrs),
                # "LIRA Value": ', '.join(lira_values_used),
                # "Fuse Name": fuse_name,
                # "Actual Value from Report": actual_value,
                # "Equation": equation,
                # "Calculated Value": result
            # })
            

        calculated_equation_output_data.append({
            "Source": source,
            "LIRA Attribute": lira_attrs,
            "LIRA Value": ', '.join(lira_values_used),
            "Fuse Name": fuse_name,
            "Actual Value from Report": actual_value,
            "HSD Info": hsd_info,
            "Equation": equation,
            "Processed Equation": processed_eqn,
            "Calculated Value": result
        })

    return calculated_equation_output_data

def evaluate_equations_mapped_result(mapped_result, fuse_dict_current, output_path, sku):
    calculated_equation_output_data = []

    # Step 1: Build a dictionary of LIRA attribute values from the JSON
    # attribute_map = {}
    # for collection in lira_json.get("collections", []):
    #     for attribute in collection.get("attributes", []):
    #         name = attribute.get("attributeName")
    #         value = attribute.get("attributeValue")
    #         if name:
    #             attribute_map[name.strip()] = value

    for fuse_name, fuse_info in fuse_dict_current.items():
        print(f"fuse_name: {fuse_name}, fuse_info: {fuse_info}")
        for item in mapped_result:
            source = item.get("Source", "Unknown")
            lira_attrs = item.get("LIRA Attribute", "Unknown")  # might contain multiple attrs
            lira_value = item.get("LIRA Value", "")
            fuse_name = item.get("Fuse Name", "Unknown").strip().split('.')[-1].lower()
            #equation = item.get("Equation", "")
            equation = item.get("Raw Equation", "") or item.get("Equation", "")
            #actual_value = fuse_dict_current.get(fuse_name, "N/A")
            actual_entry = fuse_dict_current.get(fuse_name, {})
            actual_value = actual_entry.get("value", "N/A") if isinstance(actual_entry, dict) else actual_entry
            hsd_info = actual_entry.get("comment", "") if isinstance(actual_entry, dict) else "",
            processed_eqn = equation

        break

    return 0

    # Step 2: Process each equation
    for item in mapped_result:
        source = item.get("Source", "Unknown")
        lira_attrs = item.get("LIRA Attribute", "Unknown")  # might contain multiple attrs
        lira_value = item.get("LIRA Value", "")
        fuse_name = item.get("Fuse Name", "Unknown").strip().split('.')[-1].lower()
        #equation = item.get("Equation", "")
        equation = item.get("Raw Equation", "") or item.get("Equation", "")
        #actual_value = fuse_dict_current.get(fuse_name, "N/A")
        actual_entry = fuse_dict_current.get(fuse_name, {})
        actual_value = actual_entry.get("value", "N/A") if isinstance(actual_entry, dict) else actual_entry
        hsd_info = actual_entry.get("comment", "") if isinstance(actual_entry, dict) else "",
        processed_eqn = equation
        
        # Ensure lira_attrs is a list
        if isinstance(lira_attrs, str):
            lira_attrs = [lira_attrs]

        # Clean equation for regex scan
        cleaned_eqn = str(equation).replace('"', '').replace("'", "")
        
        # Extract all LIRA attribute references case-insensitively
        matched_attrs = re.findall(r'lineitem\.([A-Za-z0-9_]+)\.value', cleaned_eqn, flags=re.IGNORECASE)
        
        # Debug print for verification
        if not matched_attrs:
            #print(f"[WARN] No LIRA attributes found in: {fuse_name} | Equation: {raw_eqn}")
            print(f"[WARN] No LIRA attributes found in: {fuse_name} | Equation: {equation}")
        else:
            print(f"[INFO] Found LIRA attributes in {fuse_name}: {matched_attrs}")
        
        # Step 3: Replace LIRA attribute placeholders in equation with actual values
        #matched_attrs = re.findall(r'lineItem\.([A-Za-z0-9_]+)\.Value', processed_eqn)
        #lira_attrs = matched_attrs  # ← this ensures correct attributes are captured   #added 1
        lira_values_used = []
        
        # Fallback if LIRA attribute field was not correctly populated
        if not matched_attrs:
            lira_attrs = ["NA"]
            lira_values_used = ["NA"]
        else: 
            # Use matched_attrs as LIRA attributes
            lira_attrs = matched_attrs
            lira_values_used = []
            for attr in matched_attrs:
                val = attribute_map.get(attr, "0")
                lira_values_used.append(f"{attr}: {val}")
                processed_eqn = processed_eqn.replace(f"lineItem.{attr}.Value", str(val))
            
        def convert_ternary(expr):
            expr = expr.replace("&&", "and").replace("||", "or")

            pattern = re.compile(r'([^\?:]+?)\?([^\?:]+?):([^\?:]+)')
            
            while '?' in expr and ':' in expr:
                    match = pattern.search(expr)
                    if not match:
                        break
                    condition = match.group(1).strip()
                    if_true = match.group(2).strip()
                    if_false = match.group(3).strip()
                    python_expr = f"({if_true} if {condition} else {if_false})"
                    expr = expr[:match.start()] + python_expr + expr[match.end():]
            return expr
   
        processed_eqn = convert_ternary(processed_eqn)
        # # Step 4: Translate .cs syntax to Python
        processed_eqn = processed_eqn.replace("Math.Round", "round") 
        processed_eqn = processed_eqn.replace("Math.Floor", "math.floor")        
        # Replace dec2hex with hex(int(round(...)))
        processed_eqn = re.sub(r'dec2hex\s*\(\s*round\((.*?)\)\s*\)', r'hex(int(round(\1)))[2:].upper()', processed_eqn)
        #processed_eqn = re.sub(r'dec2hex\s*\(\s*round\((.*?)\)\s*\)', lambda m: f"hex(int(round(float({m.group(1)}))))[2:].upper()", processed_eqn)
        processed_eqn = re.sub(r'dec2hex\s*\(\s*(.*?)\s*\)', r'hex(int(round(\1)))[2:].upper()', processed_eqn)
        #processed_eqn = re.sub(r'dec2hex\s*\(\s*(.*?)\s*\)', lambda m: f"hex(int(round(float({m.group(1)}))))[2:].upper()", processed_eqn)
        # Convert 6'hA to 0xA format
        processed_eqn = re.sub(r"(\d+)'h([0-9a-fA-F]+)", r"0x\2", processed_eqn)
        processed_eqn = re.sub(r'Convert\.ToDouble\((.*?)\)', r'float(\1)', processed_eqn)
        processed_eqn = re.sub(r'"?(\d+)\'h([0-9a-fA-F]+)"?', lambda m: hex(int(m.group(2), 16)), processed_eqn)
        processed_eqn = re.sub(r'\(?Convert\.ToDouble\((.*?)\)\)?', lambda m: f'float({m.group(1)})', processed_eqn)
        #processed_eqn = re.sub(r'dec2hex\s*\(\s*math\.floor\((.*?)\)\s*\)', r'hex(int(math.floor(\1)))[2:].upper()', processed_eqn)
        #processed_eqn = re.sub(r'dec2hex\s*\(\s*round\(math\.floor\((.*?)\)\)\s*\)', r'hex(int(math.floor(\1)))[2:].upper()', processed_eqn)
        processed_eqn = re.sub(r'dec2hex\s*\(\s*math\.floor\((.*?)\)\s*\)', r"hex(int(math.floor(\1))).replace('0x','').upper()", processed_eqn)
        processed_eqn = re.sub(r'dec2hex\s*\(\s*round\(math\.floor\((.*?)\)\)\s*\)', r"hex(int(math.floor(\1))).replace('0x','').upper()", processed_eqn)
        processed_eqn = re.sub(r'dec2hex\s*\(\s*round\((.*?)\)\s*\)',r'str(hex(int(round(\1)))).replace("0x", "").upper()',processed_eqn)

        # Step 5: Evaluate
        try:
            if processed_eqn.strip().startswith("{") and ':' in processed_eqn:
                dict_obj = eval(processed_eqn)
                key = str(lira_value).strip().upper()
                result = dict_obj.get(key, f"Key '{key}' not found")
            elif not re.search(r'[+\-*/<>=()]', processed_eqn) and not processed_eqn.strip().startswith(('lineItem', 'dec2hex')):
                result = processed_eqn
            else:
                result = eval(processed_eqn)
        except Exception as e:
            result = f"ERROR: {str(e)}"

        # Step 6: Append to results
        # if source == "DirectAssignment":
            # # Try to evaluate equation like "3'h3"
            # try:
                # evaluated_value = eval(processed_eqn)
            # except Exception as e:
                # evaluated_value = f"ERROR: {str(e)}"
        # #if not matched_attrs and not equation.strip():
            # # No LIRA attr, no equation => direct assignment case
            # calculated_equation_output_data.append({
                # "Source": source,
                # "LIRA Attribute": "No LIRA Attribute (DirectAssignment)",
                # "LIRA Value": "No LIRA Value (DirectAssignment)",
                # "Fuse Name": fuse_name,
                # "Actual Value from Report": actual_value,
                # "Equation": "",
                # "Calculated Value": actual_value
            # })
        # else:
            # calculated_equation_output_data.append({
                # "Source": source,
                # "LIRA Attribute": ', '.join(lira_attrs),
                # "LIRA Value": ', '.join(lira_values_used),
                # "Fuse Name": fuse_name,
                # "Actual Value from Report": actual_value,
                # "Equation": equation,
                # "Calculated Value": result
            # })
            

        calculated_equation_output_data.append({
            "Source": source,
            "LIRA Attribute": lira_attrs,
            "LIRA Value": ', '.join(lira_values_used),
            "Fuse Name": fuse_name,
            "Actual Value from Report": actual_value,
            "HSD Info": hsd_info,
            "Equation": equation,
            "Processed Equation": processed_eqn,
            "Calculated Value": result
        })

    return calculated_equation_output_data
    
def export_lira_evaluation_to_excel(lira_evaluation_results, output_path, sku):
    """Export LIRA mapping evaluation results to Excel"""
    import datetime
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    df = pd.DataFrame(lira_evaluation_results)
    output_file = os.path.join(output_path, f'{sku}_lira_mapping_evaluation_file5_{timestamp}.xlsx')
    df.to_excel(output_file, index=False)
    print(f"LIRA mapping evaluation results exported to: {output_file}")

def export_calculated_equation_mapping_to_excel(calculated_equation_results, output_path, sku):
    df = pd.DataFrame(calculated_equation_results)
    output_file = os.path.join(output_path, f'{sku}_calculated_fuse_equations_file4.xlsx')
    df.to_excel(output_file, index=False)
    print(f"Kavya Calculated fuse equation results exported to: {output_file}")

#def compare_excel_files(regdict1, regdict2, csv_data_dict): 
    
    # for fuse_name in regdict1.keys():
        # if fuse_name in regdict2:
            # print(f"Fuse name {fuse_name} matches in both files.")
            # utilization_method = res_dict[fuse_name]["Utilization Methods"]
                   

def main():
    print("In the main function.")
    
    # SKU name
    sku = "Q7YT"
    
    #Step 1: Parse Inputs
    regdict, fuse_dict = read_excel_fuse_file(file_path, sku)
    #print(f"\n regdict: {regdict} \n")
    #print(f"\n fuse_dict: {fuse_dict} \n")
    #return 0
    lira_json = read_lava_jason_extract(file_path)
    #print("\n Kavya LIRA JSON in main:", lira_json )
    
    equation_results= read_equations_cs_file_extract(file_path)
    print(f"\n equation_results: {equation_results} \n")
    #return 0
    #read_csv_lava_file(file_path)
    #read_excel_lava_fuse_file(file_path)
    
    # Source	LIRA Attribute	Fuse Name	Equation from .CS file
    #Step2: Export raw .cs to excel sheet ( Extracting the Fuse equation file) 
    export_cs_file_extract_to_excel(equation_results, output_path) ##==> file1
    
    #Step3: Mapping of the JSON LIRA attributes to .CS entries
    #Source	  LIRA   Attribute	LIRA Value	Fuse Name	Equation
    mapped_result = map_json_with_cs(lira_json, equation_results, output_path)
    print(f"\n mapped_result: {mapped_result}\n")
    
    
    export_map_json_cs_file_extract_to_excel(mapped_result, output_path) #==> file2
    
    #return 0
    #Step4:  Map with fuse report and export report from teh step3 
    #regdict, fuse_dict = read_excel_fuse_file(file_path, sku)
    fuse_mapped_result = map_fuse_report_with_cs(fuse_dict, mapped_result, sku)

    print(f"\n fuse_mapped_result:{fuse_mapped_result}\n")
    #print("Kavya print all fuse names available in report:", list(fuse_dict.keys()))
    #print("\n Kavya fuse mapped Results in main:", fuse_mapped_result)
    #Source	LIRA Attribute	LIRA Value	Fuse Name	Actual Value from Report	HSD Info	Equation
    export_fuse_mapping_to_excel(fuse_mapped_result, output_path,sku) #==File3
    return 0
    
    #calculated_equation_results = evaluate_equations(equation_results, lira_json, fuse_dict, output_path,sku)
    calculated_equation_results = evaluate_equations_mapped_result(mapped_result, fuse_dict, output_path,sku)

    print(f"\n calculated_equation_results:{calculated_equation_results}\n")
    #Source	LIRA Attribute	LIRA Value	Fuse Name	Actual Value from Report	HSD Info	Equation	Processed Equation	Calculated Value
    #export_calculated_equation_mapping_to_excel(calculated_equation_results, output_path,sku)
    
    

    #compare_excel_files(regdict, res_dict, csv_data_list )

    

if __name__ == "__main__":
    main()
